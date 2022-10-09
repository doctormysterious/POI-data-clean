# !/usr/bin/python
# -*- coding:utf-8 -*-

from gensim import corpora, models, similarities
from pprint import pprint

import warnings
import xlrd
import xlwt
import numpy as np
import pandas as pd
import os
import chardet
import jieba
import jieba.analyse
import codecs
import re
from openpyxl import Workbook


# import logging
# logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.INFO)
def insertOne(value1, value2, sheet):
    row = [value1, value2]
    sheet.append(row)
def get_encoding(filename):
    """
    返回文件编码格式
    """
    with open(filename,'rb') as f:
        return chardet.detect(f.read())['encoding']

def to_utf8(filename):
    """
    保存为 to_utf-8
    """
    encoding = get_encoding(filename)
    ext = os.path.splitext(filename)
    if ext[1] =='.csv':
        if 'gb' in encoding or 'GB' in encoding:
            df = pd.read_csv(filename,engine='python',encoding='GBK')
        else:
            df = pd.read_csv(filename,engine='python',encoding='utf-8')
        df.to_excel(ext[0]+'.xlsx')
        # df.to_csv(ext[0]+'_u'+'.csv')
    elif ext[1]=='.xls' or ext[1] == '.xlsx':
        if 'gb' in encoding or 'GB' in encoding:
            df = pd.read_excel(filename,encoding='GBK')
        else:
            df = pd.read_excel(filename,encoding='utf-8')
        df.to_excel(filename)
    else:
        print('only support csv, xls, xlsx format')
    print("to utf end-----------")
def batch_to_utf8(path,ext_name='csv'):
    """
    path下，后缀为 ext_name的乱码文件，批量转化为可读文件
    """
    for file in os.listdir(path):
        if os.path.splitext(file)[1]=='.'+ext_name:
            to_utf8(os.path.join(path,file))
    print("batch to utf8 end-------------")
def insertOneList(listvalue, sheet):
    sheet.append(listvalue)

def calculateType(curreader,city_name_list,dict_list):
    for k in range(len(curreader)):
        city_name = curreader['cityname'][k];
        try:
            t = city_name_list.index(city_name)
            type_name = curreader['type'][k].split(";")[0]
            if type_name not in dict_list[t]:
                for i in range(32):
                    dict_list[i][type_name] = 0
            dict_list[t][type_name] = dict_list[t][type_name] + 1
        except ValueError:
            # print(admiration_sub_list[i])
            continue
def ShowCSVTime():
    csv_file_path = 'data\\1540880282990.csv';
    csv_path = 'data\\'
    path = "D:\BaiduNetdiskDownload\gaode2022\capital\\"

    city_name_dict = {'北京市': 0, '上海市': 1, '天津市': 2, '长春市': 3, '沈阳市': 4, '呼和浩特市': 5, '石家庄市': 6, '乌鲁木齐市': 7, '兰州市': 8,
                      '西宁市': 10, '西安市': 11, '银川市': 12, '郑州市': 13, '济南市': 14, '太原市': 15, '合肥市': 16, '长沙市': 17, '武汉市': 18,
                      '南京市': 19, '成都市': 20, '贵阳市': 21, '昆明市': 22, '南宁市': 23, '拉萨市': 24, '杭州市': 25, '南昌市': 26, '广州市': 27,
                      '福州市': 28, '海口市': 29, '台北市': 30, '香港': 31, '澳门': 32}
    city_name_list = ['北京市', '上海市', '天津市', '重庆市', '长春市', '沈阳市', '哈尔滨市', '呼和浩特市', '石家庄市', '乌鲁木齐市', '兰州市', '西宁市', '西安市',
                      '银川市', '郑州市',
                      '济南市', '太原市', '合肥市', '长沙市', '武汉市', '南京市', '成都市', '贵阳市', '昆明市', '南宁市', '拉萨市', '杭州市', '南昌市', '广州市',
                      '福州市', '海口市', '台北市', '香港', '澳门']
    type_name_dict = {};
    type_name_list = ['餐饮服务', '公司企业', '汽车服务', '生活服务', '政府机构及社会团体', '购物服务', '金融保险服务', '医疗保健服务',
                      '科教文化服务', '交通设施服务', '商务住宅', '地名地址信息', '体育休闲服务', '住宿服务', '汽车维修',
                      '风景名胜', '汽车销售', '摩托车服务', '公共设施', '道路附属设施', '室内设施']
    type_dict = {}
    dict_list = []

    sampleSize = 10
    fileList = os.listdir(path)
    print("共计" + str(len(fileList)) + "个文件")
    file_num = 0;
    for filename in fileList:
        file_num = file_num + 1
        # if file_num >2:
        #     break
        print("准备读取第" + str(file_num) + "文件" + filename)
        reader = pd.read_csv(path + filename, iterator=True, chunksize=sampleSize, encoding='utf-8',
                             error_bad_lines=False,header=None,names=['a','b','address','typename','e','f','g','h','cityname','l','m','n','p','q','t'])
        # reader = pd.read_csv(path + filename, iterator=True, chunksize=sampleSize, encoding='utf-8',
        #                      error_bad_lines=False, header=None,
        #                      names=['a', 'pr', 'c', 'cityname', 'e', 'f', 'g', 'h', 'l', 't'])
        j = 0;
        # print("正在读取第" + str(file_num) + "文件" + filename)
        for b in reader:
            j = j + 1
            # print(str(j) + "---------")
            if j >= 2:
                break

            # print(b)
            for k in range((j - 1) * sampleSize, (j - 1) * sampleSize + len(b)):
                # print(k)
                c1 = b['cityname'][k];
                # c2 = b['l'][k];
                print(c1)
                c2 = b['typename'][k];
                # c2 = b['l'][k];
                print(c2)
                # print(str(c1)+":"+str(c2))
                break


        # print("已读完第" + str(file_num) + "文件" + filename)
def CalCenterCityPoiByCSV():
    csv_file_path = 'data\\1540880282990.csv';
    csv_path = 'data\\'
    path = "D:\BaiduNetdiskDownload\gaode2022\GuangDong\\"
    result_excel_location = "data\\result_data_all_2022_gd.csv"
    # reader0=pd.read_csv(csv_file_path,nrows=10,encoding='gbk')
    # city_name_dict = {'北京市': 0, '上海市': 1, '天津市': 2, '长春市': 3, '沈阳市': 4, '呼和浩特市': 5, '石家庄市': 6, '乌鲁木齐市': 7, '兰州市': 8,
    #                   '西宁市': 10, '西安市': 11, '银川市': 12, '郑州市': 13, '济南市': 14, '太原市': 15, '合肥市': 16, '长沙市': 17, '武汉市': 18,
    #                   '南京市': 19, '成都市': 20, '贵阳市': 21, '昆明市': 22, '南宁市': 23, '拉萨市': 24, '杭州市': 25, '南昌市': 26, '广州市': 27,
    #                   '福州市': 28, '海口市': 29, '台北市': 30, '香港': 31, '澳门': 32}
    # city_name_list = ['北京市', '上海市', '天津市', '重庆市', '长春市', '沈阳市', '哈尔滨市', '呼和浩特市', '石家庄市', '乌鲁木齐市', '兰州市', '西宁市', '西安市',
    #                   '银川市', '郑州市',
    #                   '济南市', '太原市', '合肥市', '长沙市', '武汉市', '南京市', '成都市', '贵阳市', '昆明市', '南宁市', '拉萨市', '杭州市', '南昌市', '广州市',
    #                   '福州市', '海口市', '台北市', '香港', '澳门']
    city_name_list = ['广州市', '韶关市', '深圳市', '珠海市', '汕头市', '佛山市', '江门市', '湛江市', '茂名市', '肇庆市', '惠州市', '梅州市',
                      '汕尾市','河源市', '阳江市','清远市', '东莞市', '潮州市', '中山市', '揭阳市', '云浮市']
    type_name_dict = {};
    type_name_list = ['餐饮服务', '公司企业', '汽车服务', '生活服务', '政府机构及社会团体', '购物服务', '金融保险服务', '医疗保健服务',
                      '科教文化服务', '交通设施服务', '商务住宅', '地名地址信息', '体育休闲服务', '住宿服务', '汽车维修',
                      '风景名胜', '汽车销售', '摩托车服务', '公共设施', '道路附属设施', '室内设施']
    type_dict = {}
    dict_list = []
    city_num=len(city_name_list)
    for i in range(city_num):
        type_dict = {}
        type_dict['city'] = city_name_list[i]
        for k in range(len(type_name_list)):
            type_dict[type_name_list[k]] = 0
        # type_dict['city'] = i
        dict_list.append(type_dict)
    sampleSize = 100000
    fileList = os.listdir(path)
    print("共计" + str(len(fileList)) + "个文件")
    file_num = 0;
    for filename in fileList:
        file_num = file_num + 1
        # if file_num >2:
        #     break
        print("准备读取第" + str(file_num) + "文件" + filename)
        reader = pd.read_csv(path + filename, iterator=True, chunksize=sampleSize, encoding='utf-8',
                             error_bad_lines=False, header=None,
                             names=['a', 'b', 'address', 'typename', 'e', 'f', 'g', 'h', 'cityname', 'l', 'm', 'n', 'p',
                                    'q', 't'],
                             usecols=['typename','cityname'])
        j = 0;
        print("正在读取第" + str(file_num) + "文件" + filename)
        for b in reader:
            j = j + 1
            print(str(j) + "---------")
            # if j <= 2:
            #     break
            #     print(b)

            for k in range((j - 1) * sampleSize, (j - 1) * sampleSize + len(b)):
                # print(k)
                city_name = b['cityname'][k];
                # print(city_name)
                # break
                try:
                    t = city_name_list.index(city_name)
                    type_name = b['typename'][k].split(";")[0]
                    if type_name.find(",")>=0:
                        type_name=type_name.split(",")[0]
                    if type_name not in dict_list[t]:
                        print(type_name)
                        for i in range(city_num):
                            dict_list[i][type_name] = 0
                    # if(j==2):
                    # print(type_name)
                    # dict_list[t][type_name]
                    dict_list[t][type_name] = dict_list[t][type_name] + 1
                except ValueError:
                    # print(admiration_sub_list[i])
                    continue
            print("已读取" + str((j - 1) * sampleSize + len(b)) + "行")
        print("已读完第" + str(file_num) + "文件" + filename)
    # else:
    #     break
    result_excel = pd.DataFrame(dict_list)
    index = 0
    for city in dict_list:
        # keys=list(city.keys())
        # values=list(city.values())
        # if index==0:
        #     result_excel['类型']=values;
        # result_excel['数值']=values;
        # result_excel=result_excel.append(result_excel,city,ignore_index=True)
        # index=index+1
        print(city)

    result_excel.to_csv(result_excel_location, encoding='gbk')
def GetBigTypeByCSV(path,result_excel_location):
    city_name_list = ['广州市', '韶关市', '深圳市', '珠海市', '汕头市', '佛山市', '江门市', '湛江市', '茂名市', '肇庆市', '惠州市', '梅州市',
                      '汕尾市', '河源市', '阳江市', '清远市', '东莞市', '潮州市', '中山市', '揭阳市', '云浮市']
    type_name_list = ['餐饮服务', '公司企业', '汽车服务', '生活服务', '政府机构及社会团体', '购物服务', '金融保险服务', '医疗保健服务',
                      '科教文化服务', '交通设施服务', '商务住宅', '地名地址信息', '体育休闲服务', '住宿服务', '汽车维修',
                      '风景名胜', '汽车销售', '摩托车服务', '公共设施', '道路附属设施', '室内设施']
    type_dict = {}
    dict_list = []
    city_num=len(city_name_list)
    for i in range(city_num):
        type_dict = {}
        type_dict['city'] = city_name_list[i]
        for k in range(len(type_name_list)):
            type_dict[type_name_list[k]] = 0
        # type_dict['city'] = i
        dict_list.append(type_dict)
    sampleSize = 100000
    fileList = os.listdir(path)
    print("共计" + str(len(fileList)) + "个文件")
    file_num = 0;
    for filename in fileList:
        file_num = file_num + 1
        # if file_num >2:
        #     break

        reader = pd.read_csv(path + filename, iterator=True, chunksize=sampleSize, encoding='utf-8',
                             error_bad_lines=False)
        j = 0;
        print("正在读取第" + str(file_num) + "文件" + filename)
        for b in reader:
            j = j + 1
            print(str(j) + "---------")
            # if j <= 2:
            #     break
            #     print(b)

            for k in range((j - 1) * sampleSize, (j - 1) * sampleSize + len(b)):
                # print(k)
                city_name = b['cityname'][k];
                try:
                    t = city_name_list.index(city_name)
                    type_name = b['type'][k].split(";")[0]
                    if type_name not in dict_list[t]:
                        for i in range(city_num):
                            dict_list[i][type_name] = 0
                    # if(j==2):
                    # print(type_name)
                    # dict_list[t][type_name]
                    dict_list[t][type_name] = dict_list[t][type_name] + 1
                except ValueError:
                    # print(admiration_sub_list[i])
                    continue
            print("已读取" + str((j - 1) * sampleSize + len(b)) + "行")
        print("已读完第" + str(file_num) + "文件" + filename)
    # else:
    #     break
    result_excel = pd.DataFrame(dict_list)
    index = 0
    for city in dict_list:
        print(city)
    result_excel.to_csv(result_excel_location, encoding='gbk')
if __name__ == '__main__':
    # path = "D:\BaiduNetdiskDownload\gaode2022\GuangDong\\"
    # result_excel_location = "data\\result_data_all_gd_2022.csv"
    # GetBigTypeByCSV(path,result_excel_location)
    # ShowCSVTime();
    # batch_to_utf8("D:\BaiduNetdiskDownload\gaode-2021\\trans\\")
    CalCenterCityPoiByCSV()

    # csv_file_path='data\\1540880282990.csv';
    # csv_path='data\\'

    # for k in range(len(reader0)):
    #     city_name=reader0['cityname'][k];
    #     try:
    #         t=city_name_list.index(city_name)
    #         type_name = reader0['type'][k].split(";")[0]
    #         if type_name not in dict_list[t]:
    #             for i in range(32):
    #                 dict_list[i][type_name] = 0
    #         dict_list[t][type_name]=dict_list[t][type_name]+1
    #     except ValueError:
    #         # print(admiration_sub_list[i])
    #         continue
    # for city in dict_list:
    #     print(city)



    # columns =reader0.dropna(axis=1).columns.tolist()

    # print(reader)


#     book = Workbook()
#     sheet = book.create_sheet("sheet" + str(2), 0)
#     sheet_cut = book.create_sheet("sheet" + str(3), 0)
#     sheets = book.sheetnames
#     f = open('2-3_data_modified.txt')
#     texts = [line.strip().split() for line in f]
# #    pprint(texts)
#     dictionary = corpora.Dictionary(texts)
#
#
#     V = len(dictionary)
#     corpus = [dictionary.doc2bow(text) for text in texts]
#     corpus_tfidf = models.TfidfModel(corpus)[corpus]
#     corpus_tfidf = corpus
#
#     print('TF-IDF:')
#    for c in corpus_tfidf:
#        print(c)

# print('\nLDA Model:')
# num_topics = 4
# lda = models.LdaModel(corpus_tfidf, num_topics=num_topics, id2word=dictionary,
#                       alpha='auto', eta='auto', minimum_probability=0.001, passes=10)
# doc_topic = [doc_t for doc_t in lda[corpus_tfidf]]
# print('Document-Topic:\n')
# #pprint(doc_topic)
# for doc_topic in lda.get_document_topics(corpus_tfidf):
#     #print(doc_topic)
#     topic=np.array(doc_topic)
#     topic_distribute=np.array(topic[:,1])
#     topic_idx =list(topic_distribute)
#     #max_value=max(topic_idx)#求主题概率的最大值
#     #max_index = topic_idx.index(max_value)#求最大值概率的主题索引
#     #insertOne(max_value,max_index,book[sheets[0]])
#     insertOneList(topic_idx,book[sheets[0]])
# for topic_id in range(num_topics):
#     print('Topic', topic_id)
#     #pprint(lda.get_topic_terms(topicid=topic_id))
#     #topic_terms =np.array(lda.show_topic(topic_id))
#     pprint(lda.show_topic(topic_id, topn=5))
#     #insertOneList(lda.show_topic(topic_id), book[sheets[1]])
# similarity = similarities.MatrixSimilarity(lda[corpus_tfidf])
# print('Similarity:')
# book.save("2-3_data_result.xlsx")
#    pprint(list(similarity))

#    hda = models.HdpModel(corpus_tfidf, id2word=dictionary)
#    topic_result = [a for a in hda[corpus_tfidf]]
#    print('\n\nUSE WITH CARE--\nHDA Model:')
#    pprint(topic_result)
#    print('HDA Topics:')
#    print(hda.print_topics(num_topics=2, num_words=5))
